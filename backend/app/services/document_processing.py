"""Document text extraction and semantic chunking.

Supports PDF, DOCX, XLSX, and image files.
Falls back to Claude Vision for scanned documents.
"""

import io
import logging
import re
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

CHARS_PER_TOKEN = 4


@dataclass
class ExtractionResult:
    text: str
    page_count: int | None = None
    is_scanned: bool = False


@dataclass
class Chunk:
    content: str
    chunk_index: int
    page_number: int | None = None
    section_ref: str | None = None
    metadata: dict = field(default_factory=lambda: {"headings": [], "keywords": []})


# ---------------------------------------------------------------------------
# Text Extraction
# ---------------------------------------------------------------------------


def extract_text(buffer: bytes, mime_type: str) -> ExtractionResult:
    """Extract text from a document buffer based on MIME type."""
    if mime_type == "application/pdf":
        return _extract_from_pdf(buffer)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        return _extract_from_docx(buffer)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_from_xlsx(buffer)
    elif mime_type.startswith("image/"):
        return ExtractionResult(text="", is_scanned=True)
    else:
        logger.warning("Unsupported MIME type: %s", mime_type)
        return ExtractionResult(text="")


def _extract_from_pdf(buffer: bytes) -> ExtractionResult:
    """Extract text from PDF using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=buffer, filetype="pdf")
        pages = []
        for page in doc:
            pages.append(page.get_text())
        text = "\n\n".join(pages)
        page_count = len(doc)
        doc.close()

        is_scanned = not text.strip() or len(text.strip()) < 100
        return ExtractionResult(text=text, page_count=page_count, is_scanned=is_scanned)
    except Exception as exc:
        logger.error("PDF extraction failed: %s", exc)
        return ExtractionResult(text="", is_scanned=True)


def _extract_from_docx(buffer: bytes) -> ExtractionResult:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document as DocxDocument

        doc = DocxDocument(io.BytesIO(buffer))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        return ExtractionResult(text=text, is_scanned=False)
    except Exception as exc:
        logger.error("DOCX extraction failed: %s", exc)
        return ExtractionResult(text="")


def _extract_from_xlsx(buffer: bytes) -> ExtractionResult:
    """Extract text from XLSX using openpyxl."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(buffer), read_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                parts.append(",".join(cells))
            parts.append("")
        wb.close()
        text = "\n".join(parts)
        return ExtractionResult(text=text, is_scanned=False)
    except Exception as exc:
        logger.error("XLSX extraction failed: %s", exc)
        return ExtractionResult(text="")


# ---------------------------------------------------------------------------
# Semantic Chunking
# ---------------------------------------------------------------------------

# Regex patterns
_PAGE_SPLIT_RE = re.compile(r"\f|(?:---\s*Page\s+\d+\s*---)", re.IGNORECASE)
_SECTION_REF_RE = re.compile(
    r"(?:Section|Division|Part)\s+[\d]+(?:[.\- ]\d+)*", re.IGNORECASE
)
_SPEC_NUMBER_RE = re.compile(r"\d{2}\s?\d{2}\s?\d{2}")
_CAPITAL_TERMS_RE = re.compile(r"\b[A-Z][A-Z]+(?:\s+[A-Z][A-Z]+){0,3}\b")


def semantic_chunk(
    text: str,
    *,
    target_tokens: int = 400,
    overlap: int = 50,
) -> list[Chunk]:
    """Split text into semantic chunks with overlap.

    Args:
        text: Full document text.
        target_tokens: Target chunk size in tokens (~4 chars/token).
        overlap: Overlap between chunks in tokens.
    """
    if not text or not text.strip():
        return []

    target_chars = target_tokens * CHARS_PER_TOKEN
    overlap_chars = overlap * CHARS_PER_TOKEN

    # Split into segments (by pages or paragraphs)
    segments = _split_segments(text)
    if not segments:
        return []

    chunks: list[Chunk] = []
    chunk_index = 0
    current_text = ""
    current_page: int | None = None

    for seg_text, seg_page in segments:
        if not seg_text.strip():
            continue

        if current_text and len(current_text) + len(seg_text) > target_chars:
            # Emit current chunk
            chunks.append(_build_chunk(current_text, chunk_index, current_page))
            chunk_index += 1

            # Apply overlap
            if overlap_chars > 0 and len(current_text) > overlap_chars:
                overlap_text = current_text[-overlap_chars:]
                current_text = overlap_text + "\n\n" + seg_text
            else:
                current_text = seg_text
            current_page = seg_page
        else:
            if current_text:
                current_text += "\n\n" + seg_text
            else:
                current_text = seg_text
            if current_page is None:
                current_page = seg_page

        # Handle oversized segments
        while len(current_text) > target_chars * 1.5:
            split_at = _find_split_point(current_text, target_chars)
            chunks.append(_build_chunk(current_text[:split_at], chunk_index, current_page))
            chunk_index += 1
            current_text = current_text[max(0, split_at - overlap_chars) :]

    # Emit final chunk
    if current_text.strip():
        chunks.append(_build_chunk(current_text, chunk_index, current_page))

    return chunks


def _split_segments(text: str) -> list[tuple[str, int | None]]:
    """Split text into segments with optional page numbers."""
    # Try page-based splitting first
    parts = _PAGE_SPLIT_RE.split(text)
    if len(parts) > 1:
        return [(part.strip(), i + 1) for i, part in enumerate(parts) if part.strip()]

    # Fall back to paragraph splitting
    paragraphs = re.split(r"\n{2,}", text)
    return [(p.strip(), None) for p in paragraphs if p.strip()]


def _find_split_point(text: str, target: int) -> int:
    """Find the best split point near the target position."""
    window = 200
    start = max(0, target - window)
    end = min(len(text), target + window)
    search_area = text[start:end]

    # Priority: paragraph break > sentence end > newline > target
    for pattern in ["\n\n", ". ", "\n"]:
        idx = search_area.rfind(pattern)
        if idx >= 0:
            return start + idx + len(pattern)

    return target


def _build_chunk(text: str, chunk_index: int, page_number: int | None) -> Chunk:
    """Build a Chunk with extracted metadata."""
    text = text.strip()

    # Extract section reference
    section_match = _SECTION_REF_RE.search(text)
    section_ref = section_match.group(0) if section_match else None

    # Extract headings (first 10 lines, length 0-100, uppercase or markdown)
    headings = []
    for line in text.split("\n")[:10]:
        line = line.strip()
        if not line or len(line) > 100:
            continue
        if line.isupper() or line.startswith("#") or re.match(r"^\d+[.\)]\s", line):
            headings.append(line)
    headings = headings[:5]

    # Extract keywords
    keywords: set[str] = set()
    # Spec numbers
    for m in _SPEC_NUMBER_RE.finditer(text):
        keywords.add(m.group(0))
    # Capital terms
    for m in _CAPITAL_TERMS_RE.finditer(text):
        term = m.group(0)
        if len(term) > 3:  # Skip short acronyms
            keywords.add(term)
    keyword_list = sorted(keywords)[:10]

    return Chunk(
        content=text,
        chunk_index=chunk_index,
        page_number=page_number,
        section_ref=section_ref,
        metadata={"headings": headings, "keywords": keyword_list},
    )
